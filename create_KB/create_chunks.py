import ast
import json
from pathlib import Path
from typing import Iterable, List, Optional

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_XLSX = REPO_ROOT / "data" / "Product_Labels.xlsx"
DEFAULT_OUTPUT_JSON = REPO_ROOT / "data" / "food_group.json"


def _as_list(value) -> List[str]:
    """
    Parse the `grouped_array` cell into a list of food names.

    Supports:
    - JSON list strings: ["a", "b"]
    - Python list strings: ['a', 'b']
    - Comma/semicolon/newline separated strings: "a, b; c"
    - Single string values
    """
    if value is None:
        return []

    if isinstance(value, (list, tuple)):
        return [str(x).strip() for x in value if str(x).strip()]

    if not isinstance(value, str):
        value = str(value)

    s = value.strip()
    if not s:
        return []

    # Try JSON
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass

        # Try python literal list
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, (list, tuple)):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass

    # Fallback split
    parts: List[str] = [s]
    for sep in [",", ";", "\n", "\r\n", "\t", "|"]:
        if sep in s:
            parts = [p.strip() for p in s.replace("\r\n", "\n").split(sep)]
            break

    cleaned = []
    for p in parts:
        p = p.strip().strip('"').strip("'").strip()
        if p:
            cleaned.append(p)
    return cleaned


def iter_food_groups(xlsx_path: Path) -> Iterable[dict]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active

    # Expect headers: parent_category, grouped_array
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError("Excel sheet is empty (no header row found).")

    header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None}
    if "parent_category" not in header_map or "grouped_array" not in header_map:
        raise ValueError(
            "Expected columns 'parent_category' and 'grouped_array' in the first row."
        )

    parent_idx = header_map["parent_category"]
    grouped_idx = header_map["grouped_array"]

    for row in rows:
        if not row:
            continue
        parent = row[parent_idx] if parent_idx < len(row) else None
        grouped = row[grouped_idx] if grouped_idx < len(row) else None

        group = "" if parent is None else str(parent).strip()
        if not group:
            continue

        for name in _as_list(grouped):
            yield {"name": name, "group": group}


def write_json(items: Iterable[dict], output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(items)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    return len(rows)


def main(input_xlsx: Optional[str] = None, output_json: Optional[str] = None) -> int:
    xlsx_path = Path(input_xlsx) if input_xlsx else DEFAULT_INPUT_XLSX
    output_path = Path(output_json) if output_json else DEFAULT_OUTPUT_JSON

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Input Excel not found: {xlsx_path}")

    count = write_json(iter_food_groups(xlsx_path), output_path)
    print(f"Wrote {count} rows to {output_path}")
    return count


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Create food chunks from Product_Labels.xlsx and write food_group.json (JSON array)"
    )
    parser.add_argument("--input", default=str(DEFAULT_INPUT_XLSX), help="Path to input .xlsx")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_JSON), help="Path to output .json")
    args = parser.parse_args()

    main(args.input, args.output)


