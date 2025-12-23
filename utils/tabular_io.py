from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook


def _decode_text(b: bytes) -> str:
    # Try utf-8-sig first (handles Excel-exported CSV with BOM), then utf-8, then latin-1 fallback
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="replace")


def read_csv_bytes(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = _decode_text(content)
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    fieldnames = list(reader.fieldnames or [])
    rows: List[Dict[str, Any]] = []
    for r in reader:
        rows.append(dict(r))
    return fieldnames, rows


def write_csv_bytes(fieldnames: List[str], rows: List[Dict[str, Any]]) -> bytes:
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in fieldnames})
    return out.getvalue().encode("utf-8-sig")


def read_xlsx_bytes(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return [], []

    fieldnames = ["" if h is None else str(h).strip() for h in headers]
    fieldnames = [f if f else f"col_{i+1}" for i, f in enumerate(fieldnames)]

    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        d: Dict[str, Any] = {}
        for i, key in enumerate(fieldnames):
            d[key] = row[i] if i < len(row) else None
        # Skip fully empty rows
        if any(v is not None and str(v).strip() != "" for v in d.values()):
            rows.append(d)

    return fieldnames, rows


def write_xlsx_bytes(fieldnames: List[str], rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for r in rows:
        ws.append([r.get(k, "") for k in fieldnames])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


